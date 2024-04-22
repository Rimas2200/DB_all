# -*- coding: utf-8 -*-
# ИФФ_1 сем 23-24 (1).xlsx
# ФЖ_1 сем  23-24 (1).xlsx

import sqlite3
import re
import openpyxl

def db_select():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS timetable(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            discipline TEXT, 
            classroom TEXT, 
            group_name TEXT,
            pair_number TEXT, 
            teacher_name TEXT,
            day_of_the_week TEXT,
            week TEXT,
            subgroup TEXT
        );
        CREATE TABLE IF NOT EXISTS groups(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT
        );
    """)
                                # discipline                    lesson_1[4::] - discipline
                                # classroom                     results
                                # group_name                    name_id
                                # pair_number                   day
                                # teacher_name                  name_with_initials
                                # day_of_the_week:              day_week
                                        # monday                    array[day_week]
                                        # tuesday                   array[day_week]
                                        # wednesday                 array[day_week]
                                        # thursday                  array[day_week]
                                        # friday                    array[day_week]
                                        # saturday                  array[day_week]
                                # week                      if "1Н" in first_two_chars:print("1Н")print(lesson_1[4::])
                                # week                      if "2Н" in first_two_chars:print("2Н")print(lesson_1[4::])
                                #subgroup                       subgroup

    # sql_ = '''    INSERT INTO
    #                     timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
    #                 VALUES
    #                     (?, ?, ?, ?, ?, ?, ?, ?)'''
    # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_week, week, subgroup))
    #                       ^^^^^^^^^^  ^^^^^^^  ^^^^^^^  ^^^  ^^^^^^^^^^^^^^^^^^  ^^^^^^^^  ^^^^  ^^^^^^^^

    db.close()


def db_start():
    db = sqlite3.connect('sql/timetable_db_test.db')
    cursor = db.cursor()
    book = openpyxl.open("расписание_2024/Расписание 2 сем ИФФ 23-24 .xlsx", read_only=True)
    sheets = book.sheetnames

    for sheet_name in sheets:
        sheet = book[sheet_name]
    #     name_ = str(sheet_name).replace(',', ' ')
    #     name__ = name_.split()
    #     lens = len(name__)
        array = [" ", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", ]; i = 0; n = 2; f = 0; colors = 0; day_week = 1;
        name_with_initials = ''
        array_week = 1
        substring = "ОНЛАЙН-КУРС"
        day = None
        subgroup = ''
        while i != 200:
            try:
                name = sheet.cell(row=11, column=3 + f).value
                name_id = name.replace(' ', '')

                color = 0
                try:
                    cell = sheet.cell(row=12 + i, column=2)
                    fill = cell.fill
                    color = fill.fgColor.rgb if fill.fgColor.rgb is not None else 'No Fill'
                    # print(color)
                except: pass
                if sheet.cell(row=12 + i, column=2).value is not None:
                    day = str(sheet.cell(row=12 + i, column=2).value)
                else:
                    day = str(sheet.cell(row=11 + i, column=2).value)
                if str(color) != "FF00B0F0":
                    if sheet.cell(row=12+i, column=3+f).value is not None:
                        string = sheet.cell(row=12 + i, column=3+f).value
                        if sheet.cell(row=12 + i, column=4 + f).value is not None:
                            subgroup = "первая подгруппа"
                        else:
                            subgroup = " "
                        try:
                            lessons = string.split("\n")
                            first_two_chars = string[:4]
                            result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[1])
                            if result:
                                teacher_name = result.group()
                                name = teacher_name.title()
                                name_with_initials = name[:-2] + ". " + name[-1] + "."
                            strings = lessons[0]
                            if substring is strings:
                                matches_aud = substring
                            else:
                                matches = lessons[1]
                                pattern = r"\b(\d{3})\b" # для аудитории
                                matches_aud = re.findall(pattern, matches)
                                matches_aud = 'АУД. ' + matches_aud[0]

                            if first_two_chars == "2 Н.":
                                para = lessons[0]
                                week = '2Н'
                                discipline = para[5:]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                                # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                            if first_two_chars == "1 Н.":
                                para = lessons[0]
                                week = '1Н'
                                discipline = para[5:]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                            if "1 Н." not in first_two_chars and "2 Н." not in first_two_chars:
                                week = '0'
                                discipline = lessons[0]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                        except: pass

                    if sheet.cell(row=12+i, column=4+f).value is not None:
                        string = sheet.cell(row=12 + i, column=4+f).value
                        try:
                            lessons = string.split("\n")
                            first_two_chars = string[:4]
                            result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[1])
                            if result:
                                teacher_name = result.group()
                                name = teacher_name.title()
                                name_with_initials = name[:-2] + ". " + name[-1] + "."
                            strings = lessons[0]
                            if substring is strings:
                                matches_aud = substring
                            else:
                                matches = lessons[1]
                                pattern = r"\b(\d{3})\b" # для аудитории
                                matches_aud = re.findall(pattern, matches)
                                matches_aud = 'АУД. ' + matches_aud[0]

                            if first_two_chars == "2 Н.":
                                para = lessons[0]
                                week = '2Н'
                                discipline = para[5:]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, "вторая подгруппа")
                                # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, "вторая подгруппа"))
                            if first_two_chars == "1 Н.":
                                para = lessons[0]
                                week = '1Н'
                                discipline = para[5:]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, "вторая подгруппа")
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, "вторая подгруппа"))
                            if "1 Н." not in first_two_chars and "2 Н." not in first_two_chars:
                                week = '0'
                                discipline = lessons[0]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, "вторая подгруппа")
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, "вторая подгруппа"))
                        except: pass



                    i += 1
                else:
                    colors += 1
                    n += 1
                    i += 1
                    array_week += 1
                    print("----------------------------------------------------------")
                    color = 0
                if colors == 6:
                    f += 2; i = 0; j = 0; n = 2; colors = 0; string = str(""); string_2 = str(""); day = "None"; array_week = 1;
            except:
                f += 2
                i = 0
                j = 0
                n = 2
                colors = 0
                string = str("")
                string_2 = str("")
                day = "None"
                array_week = 1
                break
    db.commit()

def test():
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    def get_cell_color(file_path, sheet_name, cell_address):
        # Загрузка книги Excel
        workbook = load_workbook(file_path)
        # Выбор листа по имени
        sheet = workbook[sheet_name]
        # Получение ячейки по адресу
        cell = sheet[cell_address]
        # Получение цвета заливки ячейки
        fill = cell.fill
        if isinstance(fill, PatternFill):
            cell_color = fill.fgColor.rgb
        else:
            cell_color = None
        return cell_color

    # Пример использования
    file_path = 'ФЖ_1 сем  23-24 (1).xlsx'
    sheet_name = 'ФЖ1'
    cell_address = 'C12'

    color = get_cell_color(file_path, sheet_name, cell_address)
    print(f'{cell_address}: {color}')

def db_f():
    db = sqlite3.connect('sql/timetable_db_test.db')
    cursor = db.cursor()
    cursor.execute('''SELECT * FROM timetable''')
    results = cursor.fetchall()
    for row in results:
        print(row)

db_start()
# db_select()
# test()
db_f()


