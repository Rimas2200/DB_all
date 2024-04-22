# юристы-1_23-24.xlsx
# -*- coding: utf-8 -*-

import re
import sqlite3
import openpyxl

def db_select():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS timetable(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            discipline TEXT, 
            classroom TEXT, 
            group_name TEXT,
            pair_number TEXT, 
            teacher_name TEXT,
            day_of_the_week TEXT,
            week TEXT,
            subgroup TEXT
        );
        CREATE TABLE IF NOT EXISTS groups(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT
        );
    """)
                                # discipline                    lesson_1[4::] - discipline
                                # classroom                     results
                                # group_name                    name_id
                                # pair_number                   day
                                # teacher_name                  name_with_initials
                                # day_of_the_week:              day_week
                                        # monday                    array[day_week]
                                        # tuesday                   array[day_week]
                                        # wednesday                 array[day_week]
                                        # thursday                  array[day_week]
                                        # friday                    array[day_week]
                                        # saturday                  array[day_week]
                                # week                      if "1Н" in first_two_chars:print("1Н")print(lesson_1[4::])
                                # week                      if "2Н" in first_two_chars:print("2Н")print(lesson_1[4::])
                                #subgroup                       subgroup

    # sql_ = '''    INSERT INTO
    #                     timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
    #                 VALUES
    #                     (?, ?, ?, ?, ?, ?, ?, ?)'''
    # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_week, week, subgroup))
    #                       ^^^^^^^^^^  ^^^^^^^  ^^^^^^^  ^^^  ^^^^^^^^^^^^^^^^^^  ^^^^^^^^  ^^^^  ^^^^^^^^

    db.close()

def db_start():
    discipline = None
    results = None
    name_with_initials = None
    week = None
    name_with_initials_1 = None
    rows = 15
    columns = 5
    columns_pars = 16
    substring = "(ОНЛАЙН-КУРС)"

    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    book = openpyxl.open("юристы-1_23-24.xlsx", read_only=True)  # открытие на чтение
    sheets = book.sheetnames
    for sheet_name in sheets:
        sheet = book[sheet_name]
        name_ = str(sheet_name).replace(',', ' ')
        name__ = name_.split()
        lens = len(name__)
        array = [" ", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", ]; i = 0; n = 2; f = 0; colors = 0; day_week = 1;
        colors = 1
        day = "None"
        discipline = None
        results = None
        name_with_initials = "Вакансия"
        name_with_initials_1 = "Вакансия"
        week = None
        color = ''
        f = 0
        while i != 250:
            subgroup = ''
            try:
                cell = sheet.cell(row=11 + i, column=2)
                fill = cell.fill
                color = fill.fgColor.rgb if fill.fgColor.rgb is not None else 'No Fill'
                # print(color)
            except:
                pass
            lessons = []
            if str(color) != "FF00B0F0":
                day = sheet.cell(row=11 + i, column=2).value
                try:
                    for j in range(0, 3):
                        para_1 = sheet.cell(row=11 + i + j, column=3 + f).value
                        para = sheet.cell(row=11 + i + j, column=3 + f)
                        if sheet.cell(row=11 + i + j, column=5 + f).value is not None:
                            subgroup = "первая подгруппа"
                        if sheet.cell(row=11 + i + j, column=3 + f).value is not None:
                            lessons.append(sheet.cell(row=11 + i + j, column=3 + f).value)
                        else:
                            lessons.append(None)

                        if sheet.cell(row=10, column=3 + f).value is not None:
                            name_id = sheet.cell(row=10, column=3 + f).value
                            name_id = str(name_id).replace(' ', '')
                    i+=3
                    discipline = str(lessons[0])
                    first_two_chars = discipline[:3]
                    first = str(lessons[1])
                    first_two = first[:3]

                    if "1Н " not in first_two_chars and "2Н " not in first_two_chars and "2Н " not in first_two:
                        discipline = lessons[0]
                        matches_aud = lessons[1]
                        name_with_initials = lessons[2]
                        week = "0"
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(name_with_initials))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))

                    discipline = str(lessons[1])
                    first_two_chars = discipline[:3]

                    if first_two_chars == "2Н " and lessons[0] is None:
                        discipline = lessons[1]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)
                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass
                        name_with_initials = lessons[2]
                        week = "2Н"
                        discipline = str(discipline).replace('2Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(name_with_initials))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_,(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))

                    discipline = str(lessons[0])
                    first_two_chars = discipline[:3]
                    discipline_ = str(lessons[2])
                    first_two_chars_ = discipline_[:3]
                    if first_two_chars == '1Н ' and first_two_chars_ == '2Н ':
                        discipline = lessons[0]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)

                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass
                        name_with_initials = lessons[1]
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(name_with_initials))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        week = '1Н'
                        discipline = str(discipline).replace('1Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))

                        discipline = lessons[2]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)
                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(discipline))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        week = '2Н'
                        discipline = str(discipline).replace('2Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        pattern = r"доцент\s+\w+\s+\w+\."
                        discipline = re.sub(pattern, "", discipline)
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))

                    discipline = str(lessons[0])
                    discipline_ = str(lessons[1])
                    first_two_chars = discipline[:3]
                    first_two_chars_ = discipline_[:3]
                    if "1Н " in first_two_chars and "2Н " in first_two_chars_:
                        discipline = lessons[0]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)
                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(discipline))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        week = '1Н'
                        discipline = str(discipline).replace('1Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        pattern = r"доцент\s+\w+\s+\w+\."
                        discipline = re.sub(pattern, "", discipline)
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))

                        discipline = lessons[1]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)
                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass
                        matches_aud = matches_aud[1]
                        name_with_initials = lessons[2]
                        week = "2Н"
                        discipline = str(discipline).replace('2Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(name_with_initials))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))


                    discipline = str(lessons[0])
                    discipline_ = str(lessons[2])
                    first_two_chars = discipline[:3]
                    if "1Н " in first_two_chars and lessons[2] is None:
                        discipline = lessons[0]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)
                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(lessons[1]))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        week = '1Н'
                        discipline = str(discipline).replace('1Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        pattern = r"доцент\s+\w+\s+\w+\."
                        discipline = re.sub(pattern, "", discipline)
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))

                        discipline = lessons[1]
                        match = re.search(r"\b([А-Я]-\d+)\b", discipline)
                        if match is not None:
                            matches_aud = match.group(1)
                        else:
                            matches_aud = re.findall(r'\d+', str(discipline))
                            try: matches_aud = matches_aud[1]
                            except: pass

                        name_with_initials = lessons[2]
                        week = "2Н"
                        discipline = str(discipline).replace('2Н ', '')
                        discipline = str(discipline).replace(str(matches_aud), '')
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", str(name_with_initials))
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        day_of_the_week = array[colors]
                        if discipline is not None and matches_aud is not None and name_id is not None and day is not None and name_with_initials is not None and week is not None:
                            print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                            sql_ = '''    INSERT INTO
                                         timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                     VALUES
                                         (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                except:pass


            else:
                colors += 1
                i-=2
                j=0
                lessons = []
            if colors == 7:
                f += 4;
                i = 0;
                j = 0;
                n = 2;
                colors = 1;
                lens -= 1;
                string = str("");
                string_2 = str("");
                day = "None";
                array_week = 1;
            if lens == 0:
                break
    db.commit()
    cursor.execute('''SELECT * FROM timetable''')
    results = cursor.fetchall()
    for row in results:
        print(row)


def db_f():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    cursor.execute('''SELECT * FROM timetable''')
    results = cursor.fetchall()
    for row in results:
        print(row)

# db_select()
db_start()
db_f()
