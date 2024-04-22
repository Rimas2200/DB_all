# -*- coding: utf-8 -*-
# ФЛИП_1_СЕМ_2023-2024 (1).xlsx

import re
import sqlite3
import openpyxl

def db_select():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS timetable(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            discipline TEXT, 
            classroom TEXT, 
            group_name TEXT,
            pair_number TEXT, 
            teacher_name TEXT,
            day_of_the_week TEXT,
            week TEXT,
            subgroup TEXT
        );
        CREATE TABLE IF NOT EXISTS groups(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT
        );
    """)
                                # discipline                    lesson_1[4::] - discipline
                                # classroom                     results
                                # group_name                    name_id
                                # pair_number                   day
                                # teacher_name                  name_with_initials
                                # day_of_the_week:              day_week
                                        # monday                    array[day_week]
                                        # tuesday                   array[day_week]
                                        # wednesday                 array[day_week]
                                        # thursday                  array[day_week]
                                        # friday                    array[day_week]
                                        # saturday                  array[day_week]
                                # week                      if "1Н" in first_two_chars:print("1Н")print(lesson_1[4::])
                                # week                      if "2Н" in first_two_chars:print("2Н")print(lesson_1[4::])
                                #subgroup                       subgroup

    # sql_ = '''    INSERT INTO
    #                     timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
    #                 VALUES
    #                     (?, ?, ?, ?, ?, ?, ?, ?)'''
    # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_week, week, subgroup))
    #                       ^^^^^^^^^^  ^^^^^^^  ^^^^^^^  ^^^  ^^^^^^^^^^^^^^^^^^  ^^^^^^^^  ^^^^  ^^^^^^^^

    db.close()

def db_start():
    db = sqlite3.connect('sql/timetable_db_test.db')
    cursor = db.cursor()
    book = openpyxl.open("расписание_2024/ФЛИП_2_СЕМ_2023-2024.xlsx", read_only=True)
    sheets = book.sheetnames

    for sheet_name in sheets:
        sheet = book[sheet_name]
    #     name_ = str(sheet_name).replace(',', ' ')
    #     name__ = name_.split()
    #     lens = len(name__)
        array = [" ", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", ]; i = 0; n = 2; f = 0; colors = 0; day_week = 1;
        name_with_initials = ''
        array_week = 1
        substring = "ОНЛАЙН-КУРС"
        day = None
        subgroup = ''
        while i != 200:
            try:
                name = sheet.cell(row=12, column=3 + f).value
                name_id = name.replace(' ', '')

                color = 0
                try:
                    cell = sheet.cell(row=14 + i, column=2)
                    fill = cell.fill
                    color = fill.fgColor.rgb if fill.fgColor.rgb is not None else 'No Fill'
                except: pass
                if sheet.cell(row=14 + i, column=2).value is not None:
                    day = str(sheet.cell(row=14 + i, column=2).value)
                else:
                    day = str(sheet.cell(row=12 + i, column=2).value)
                if str(color) != "FF00B0F0":
                    if sheet.cell(row=14+i, column=3+f).value is not None:
                        string = sheet.cell(row=14 + i, column=3+f).value

                        try:
                            lessons = string.split("\n")
                            first_two_chars = string[:4]
                            if "1 Н." not in first_two_chars and "2 Н." not in first_two_chars:
                                less = lessons[1]+lessons[2]
                            if "1 Н." in first_two_chars or "2 Н." in first_two_chars:
                                less = lessons[1]
                            result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", less)
                            if result:
                                teacher_name = result.group()
                                name = teacher_name.title()
                                name_with_initials = name[:-2] + ". " + name[-1] + "."
                            strings = lessons[0]
                            try:
                                if substring is strings:
                                    matches_aud = substring
                                else:
                                    matches = less
                                    pattern = r"\b(\d{3})\b" # для аудитории
                                    matches_aud = re.findall(pattern, matches)
                                    matches_aud = 'АУД. ' + matches_aud[0]
                            except:
                                split_text = less.split(",")  # Разделение строки по запятой
                                auditorium = split_text[-1].strip()
                                matches_aud = 'АУД. ' + auditorium
                            if first_two_chars == "2 Н.":
                                para = lessons[0]
                                week = '2Н'
                                discipline = para[5:]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                                # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                            if first_two_chars == "1 Н.":
                                para = lessons[0]
                                week = '1Н'
                                discipline = para[5:]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                            if "1 Н." not in first_two_chars and "2 Н." not in first_two_chars:
                                week = '0'
                                discipline = lessons[0]
                                day_of_the_week = array[array_week]
                                print(discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                                sql_ = '''    INSERT INTO
                                                        timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                    VALUES
                                                        (?, ?, ?, ?, ?, ?, ?, ?)'''
                                cursor.execute(sql_, (discipline, matches_aud, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                        except: pass



                    i += 1
                else:
                    colors += 1
                    n += 1
                    i += 1
                    array_week += 1
                    print("----------------------------------------------------------")
                    color = 0
                if colors == 6:
                    f += 1; i = 0; j = 0; n = 2; colors = 0; string = str(""); string_2 = str(""); day = "None"; array_week = 1;
            except:
                f += 2
                i = 0
                j = 0
                n = 2
                colors = 0
                string = str("")
                string_2 = str("")
                day = "None"
                array_week = 1
                break
    db.commit()
def db_f():
    db = sqlite3.connect('sql/timetable_db_test.db')
    cursor = db.cursor()
    cursor.execute('''SELECT * FROM timetable''')
    results = cursor.fetchall()
    for row in results:
        print(row)


def db_del():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()

    cursor.execute('''SELECT * FROM timetable WHERE id >= 1642''')
    results = cursor.fetchall()
    for row in results:
        cursor.execute('''DELETE FROM timetable WHERE id = ?''', (row[0],))
        print("Удалена запись с ID", row[0])
    db.commit()
    db.close()

def db_men():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()

    # Выполнение запроса к базе данных для получения данных из столбца 'group_name'
    cursor.execute("SELECT group_name FROM timetable")

    # Извлечение всех данных из результата запроса
    results = cursor.fetchall()

    # Создание словаря для хранения данных в требуемом формате
    group_data = {
        'МК': [],
        'МН': [],
        'МП': [],
        'МПМ': [],
        'МТ': [],
        'МТМ': [],
        'МММ': []
    }

    for row in results:
        group_name = row[0]

        # Проверка принадлежности группы к определенной категории и добавление ее в соответствующий список
        if group_name.startswith('МК-') and group_name not in group_data['МК']:
            group_data['МК'].append(group_name)
        elif group_name.startswith('МН-') and group_name not in group_data['МН']:
            group_data['МН'].append(group_name)
        elif group_name.startswith('МП-') and group_name not in group_data['МП']:
            group_data['МП'].append(group_name)
        elif group_name.startswith('МПМ-') and group_name not in group_data['МПМ']:
            group_data['МПМ'].append(group_name)
        elif group_name.startswith('МТ-') and group_name not in group_data['МТ']:
            group_data['МТ'].append(group_name)
        elif group_name.startswith('МТМ-') and group_name not in group_data['МТМ']:
            group_data['МТМ'].append(group_name)
        elif group_name.startswith('МММ-') and group_name not in group_data['МММ']:
            group_data['МММ'].append(group_name)

    # Вывод данных в требуемом формате
    for group, names in group_data.items():
        print(f"'{group}': {names},")

db_start()
db_f()
# db_men()
# db_del()