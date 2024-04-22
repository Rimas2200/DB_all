# -*- coding: utf-8 -*-

import re
import sqlite3
import openpyxl

def db_select():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS timetable(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            discipline TEXT, 
            classroom TEXT, 
            group_name TEXT,
            pair_number TEXT, 
            teacher_name TEXT,
            day_of_the_week TEXT,
            week TEXT,
            subgroup TEXT
        );
        CREATE TABLE IF NOT EXISTS groups(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT
        );
    """)
                                # discipline                    lesson_1[4::] - discipline
                                # classroom                     results
                                # group_name                    name_id
                                # pair_number                   day
                                # teacher_name                  name_with_initials
                                # day_of_the_week:              day_week
                                        # monday                    array[day_week]
                                        # tuesday                   array[day_week]
                                        # wednesday                 array[day_week]
                                        # thursday                  array[day_week]
                                        # friday                    array[day_week]
                                        # saturday                  array[day_week]
                                # week                      if "1Н" in first_two_chars:print("1Н")print(lesson_1[4::])
                                # week                      if "2Н" in first_two_chars:print("2Н")print(lesson_1[4::])
                                #subgroup                       subgroup

    # sql_ = '''    INSERT INTO
    #                     timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
    #                 VALUES
    #                     (?, ?, ?, ?, ?, ?, ?, ?)'''
    # cursor.execute(sql_, (discipline, results, name_id, day, name_with_initials, day_week, week, subgroup))
    #                       ^^^^^^^^^^  ^^^^^^^  ^^^^^^^  ^^^  ^^^^^^^^^^^^^^^^^^  ^^^^^^^^  ^^^^  ^^^^^^^^

    db.close()

def db_start():
    discipline = None
    results = None
    name_with_initials = None
    week = None
    name_with_initials_1 = None
    rows = 15
    columns = 5
    columns_pars = 16
    substring = "(ОНЛАЙН-КУРС)"

    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    book = openpyxl.open("расписание_2024/БФ_1 сем 23-24 (1).xlsx", read_only=True)  # открытие на чтение
    sheets = book.sheetnames
    for sheet_name in sheets:
        sheet = book[sheet_name]
        name_ = str(sheet_name).replace(',', ' ')
        name__ = name_.split()
        lens = len(name__)
        array = [" ", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", ]; i = 0; n = 2; f = 0; colors = 0; day_week = 1;

        day = "None"
        discipline = None
        results = None
        name_with_initials = "Вакансия"
        name_with_initials_1 = "Вакансия"
        week = None
        while i != 200:
            string = str("")
            string_2 = str("")
            for j in range(0, 4):
                try:
                    cell = sheet.cell(row=columns_pars + i, column=4)
                    fill = cell.fill
                    color = fill.fgColor.rgb if fill.fgColor.rgb is not None else 'No Fill'
                except:
                    pass
                if str(color) != "FF00B0F0":
                    para_1 = sheet.cell(row=columns_pars + i + j, column=5 + f).value
                    para_2 = sheet.cell(row=columns_pars + i + j, column=6 + f).value
                    if para_1 is not None:
                        day = str(sheet.cell(row=columns_pars + i, column=4).value)
                        para_1 = sheet.cell(row=columns_pars + i + j, column=5 + f).value
                        if para_1 != "None":
                            string = string + str(para_1) + "\n"
                    if para_2 is not None:
                        day = str(sheet.cell(row=columns_pars + i, column=4).value)
                        para_2 = sheet.cell(row=columns_pars + i + j, column=6 + f).value
                        if para_2 != "None":
                            string_2 = string_2 + str(para_2) + "\n"
                else:
                    colors += 1; n += 1; i -= 3
                    day_week += 1
                    if name is not None:
                        name = sheet.cell(row=rows, column=columns + f).value
                        print(sheet.cell(row=rows, column=columns + f))
                        name_id = name.replace(' ', '-')
                    else:
                        break
            # print(name_id)
            if day != "None":
                name = sheet.cell(row=rows, column=columns + f).value
                name_id = name.replace(' ', '-')
                if string_2 and string:
                    lessons = string.split("\n")
                    result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[1])
                    lesson_1 = lessons[0]
                    subgroup = 'первая подгруппа'
                    first_two_chars = string[:2]
                    #print("1", first_two_chars)
                    if "1Н" in first_two_chars:
                        week = "1Н"
                        discipline = (lesson_1[3::])
                    if "2Н" in first_two_chars:
                        week = "2Н"
                        discipline = (lesson_1[3::])
                    if "1Н" not in first_two_chars and "2Н" not in first_two_chars:
                        discipline = lesson_1
                    if result:
                        teacher_name = result.group()
                        name = teacher_name.title()
                        name_with_initials = name[:-2] + ". " + name[-1] + "."
                    aud = re.search(r".*?(АУД\.\s\d+)$", string)
                    if aud is None:
                        resultt = string.split(", ")[-1]
                        if "АУД" in resultt:
                            extracted_string = resultt.strip()
                            results = extracted_string
                    else:
                        results = aud.group(1)
                    day_of_the_week = array[day_week]
                    sql_ = '''    INSERT INTO
                                                                    timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                                VALUES
                                                                    (?, ?, ?, ?, ?, ?, ?, ?)'''
                    cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                    print(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup)

                    if lessons[2]:
                        lessons = string.split("\n")
                        string = lessons[2] + lessons[3]
                        lesson_1 = lessons[2]
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[3])
                        subgroup = 'первая подгруппа'
                        discipline = ""
                        first_two_chars = string[:2]
                        #print("2", first_two_chars)
                        if "1Н" in first_two_chars:
                            week = "1Н"
                            discipline = (lesson_1[3::])
                        if "2Н" in first_two_chars:
                            week = "2Н"
                            discipline = (lesson_1[3::])
                        if "1Н" not in first_two_chars and "2Н" not in first_two_chars:
                            week = "0"
                            discipline = lesson_1
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        aud = re.search(r".*?(АУД\.\s\d+)$", string)
                        if aud is None:
                            resultt = string.split(", ")[-1]
                            if "АУД" in resultt:
                                extracted_string = resultt.strip()
                                results = extracted_string
                        else:
                            results = aud.group(1)
                            # print(results)
                        day_of_the_week = array[day_week]
                        sql_ = '''    INSERT INTO
                                                timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                            VALUES
                                                (?, ?, ?, ?, ?, ?, ?, ?)'''
                        cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                        print(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                    lessons = string_2.split("\n")
                    lesson_2 = lessons[0]
                    result_1 = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[1])
                    subgroup = 'вторая подгруппа'
                    first_two_chars2 = string_2[:2]
                    #print("3", first_two_chars2)
                    if "1Н" in first_two_chars2:
                        week = "1Н"
                        discipline = (lesson_2[3::])
                    if "2Н" in first_two_chars2:
                        week = "2Н"
                        discipline = (lesson_2[3::])
                    if "1Н" not in first_two_chars2 and "2Н" not in first_two_chars2:
                        week = "0"
                        discipline = lesson_2
                    if result_1:
                        teacher_name_1 = result_1.group()
                        name_1 = teacher_name_1.title()
                        name_with_initials_1 = name_1[:-2] + ". " + name_1[-1] + "."
                    aud = re.search(r".*?(АУД\.\s\d+)$", lessons[1])
                    # print(aud, "=======================")
                    if aud is None:
                        resultt = string_2.split(", ")[-1]
                        if "АУД" in resultt:
                            extracted_string = resultt.strip()
                            results = extracted_string
                    else:
                        results = aud.group(1)
                    day_of_the_week = array[day_week]
                    sql_ = '''    INSERT INTO
                                            timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                        VALUES
                                            (?, ?, ?, ?, ?, ?, ?, ?)'''
                    cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials_1, day_of_the_week, week, subgroup))
                    print(discipline, results, name_id, day, name_with_initials_1, day_of_the_week, week, subgroup)
                    if lessons[2]:
                        lessons = string_2.split("\n")
                        string = lessons[2] + lessons[3]
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[3])
                        lesson_1 = lessons[2]
                        subgroup = 'вторая подгруппа'
                        first_two_chars = string[:2]
                        #print("4", first_two_chars)
                        if "1Н" in first_two_chars:
                            week = "1Н"
                            discipline = (lesson_1[3::])
                        if "2Н" in first_two_chars:
                            week = "2Н"
                            discipline = (lesson_1[3::])
                        if "1Н" not in first_two_chars and "2Н" not in first_two_chars:
                            week = "0"
                            discipline = lesson_1
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        aud = re.search(r".*?(АУД\.\s\d+)$", string)
                        if aud is None:
                            resultt = string.split(", ")[-1]
                            if "АУД" in resultt:
                                extracted_string = resultt.strip()
                                results = extracted_string
                        else:
                            results = aud.group(1)
                        day_of_the_week = array[day_week]
                        sql_ = '''    INSERT INTO
                                                timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                            VALUES
                                                (?, ?, ?, ?, ?, ?, ?, ?)'''
                        cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                        print(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup)

                if string and not string_2:
                    subgroup = " "
                    name_with_initials = "Вакансия"
                    name_with_initials_1 = "Вакансия"
                    lessons = string.split("\n")
                    lesson_1 = lessons[0]
                    result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[1])
                    first_two_chars = string[:2]
                    #print("5", first_two_chars)
                    if "1Н" in first_two_chars:
                        week = "1Н"
                        discipline = (lesson_1[3::])
                    if "2Н" in first_two_chars:
                        week = "2Н"
                        discipline = (lesson_1[3::])
                    if "1Н" not in first_two_chars and "2Н" not in first_two_chars:
                        week = "0"
                        discipline = lesson_1
                    if result:
                        teacher_name = result.group()
                        name = teacher_name.title()
                        name_with_initials = name[:-2] + ". " + name[-1] + "."
                    if substring in string:
                        # print(lessons[0])
                        results = "Онлайн-курс"
                    else:

                        aud = re.search(r".*?(АУД\.\s\d+)$", string)
                        if aud is None:
                            resultt = string.split(", ")[-1]
                            if "АУД" in resultt:
                                extracted_string = resultt.strip()
                                results = extracted_string
                        else:
                            results = aud.group(1)
                    day_of_the_week = array[day_week]
                    sql_ = '''    INSERT INTO
                                            timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                        VALUES
                                            (?, ?, ?, ?, ?, ?, ?, ?)'''
                    cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                    print(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                    try:
                        if lessons[2]:
                            subgroup = " "
                            name_with_initials = " "
                            name_with_initials_1 = " "
                            lessons = string.split("\n")
                            string = lessons[2] + lessons[3]
                            lesson_1 = lessons[2]
                            result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[3])
                            first_two_chars = string[:2]
                            #print("6", first_two_chars)
                            if "1Н" in first_two_chars:
                                week = "1Н"
                                discipline = (lesson_1[3::])
                            if "2Н" in first_two_chars:
                                week = "2Н"
                                discipline = (lesson_1[3::])
                            if "1Н" not in first_two_chars and "2Н" not in first_two_chars:
                                week = "0"
                                discipline = lesson_1
                            if result:
                                teacher_name = result.group()
                                name = teacher_name.title()
                                name_with_initials = name[:-2] + ". " + name[-1] + "."
                            if substring in string:
                                # print(lessons[0])
                                results = "Онлайн-курс"
                            else:

                                aud = re.search(r".*?(АУД\.\s\d+)$", string)
                                if aud is None:
                                    resultt = string.split(", ")[-1]
                                    if "АУД" in resultt:
                                        extracted_string = resultt.strip()
                                        results = extracted_string
                                else:
                                    results = aud.group(1)
                            day_of_the_week = array[day_week]
                            sql_ = '''    INSERT INTO
                                                    timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                                VALUES
                                                    (?, ?, ?, ?, ?, ?, ?, ?)'''
                            cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                            print(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup)
                    except:
                        print(day)
                        print(name_id)
                        print(array[day_week])
                if string_2 and not string:
                    result_1 = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", string_2)
                    lessons = string_2.split("\n")
                    lesson_2 = lessons[0]
                    subgroup = 'вторая подгруппа'
                    first_two_chars2 = string_2[:2]
                    #print("7", first_two_chars2)
                    if "1Н" in first_two_chars2:
                        week = "1Н"
                        discipline = (lesson_2[3::])
                    if "2Н" in first_two_chars2:
                        week = "2Н"
                        discipline = (lesson_2[3::])
                    if "1Н" not in first_two_chars2 and "2Н" not in first_two_chars2:
                        week = "0"
                        discipline = lesson_2
                    if result_1:
                        teacher_name_1 = result_1.group()
                        name_1 = teacher_name_1.title()
                        name_with_initials_1 = name_1[:-2] + ". " + name_1[-1] + "."
                    aud = re.search(r".*?(АУД\.\s\d+)$", string_2)
                    if aud is None:
                        resultt = string.split(", ")[-1]
                        if "АУД" in resultt:
                            extracted_string = resultt.strip()
                            results = extracted_string
                    else:
                        results = aud.group(1)
                    day_of_the_week = array[day_week]
                    sql_ = '''    INSERT INTO
                                            timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                        VALUES
                                            (?, ?, ?, ?, ?, ?, ?, ?)'''
                    cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials_1, day_of_the_week, week, subgroup))
                    print(discipline, results, name_id, day, name_with_initials_1, day_of_the_week, week, subgroup)
                    if lessons[2]:
                        subgroup = " "
                        lessons = string_2.split("\n")
                        print(string, lessons)
                        string = lessons[2] + lessons[3]
                        lesson_1 = lessons[2]
                        result = re.search(r"\b([А-Я]\w+\.? [А-Я]\.?[А-Я]?)\b", lessons[3])
                        first_two_chars = string[:2]
                        #print("8", first_two_chars)
                        if "1Н" in first_two_chars:
                            week = "1Н"
                            discipline = (lesson_1[3::])
                        if "2Н" in first_two_chars:
                            week = "2Н"
                            discipline = (lesson_1[3::])
                        if "1Н" not in first_two_chars and "2Н" not in first_two_chars:
                            week = "0"
                            discipline = lesson_1
                        if result:
                            teacher_name = result.group()
                            name = teacher_name.title()
                            name_with_initials = name[:-2] + ". " + name[-1] + "."
                        aud = re.search(r".*?(АУД\.\s\d+)$", string)
                        if aud is None:
                            resultt = string.split(", ")[-1]
                            if "АУД" in resultt:
                                extracted_string = resultt.strip()
                                results = extracted_string
                        else:
                            results = aud.group(1)
                        day_of_the_week = array[day_week]
                        sql_ = '''    INSERT INTO
                                                timetable(discipline, classroom, group_name, pair_number, teacher_name, day_of_the_week, week, subgroup)
                                            VALUES
                                                (?, ?, ?, ?, ?, ?, ?, ?)'''
                        cursor.execute(sql_,(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup))
                        print(discipline, results, name_id, day, name_with_initials, day_of_the_week, week, subgroup)

            i += 4
            if colors == 6:
                print("----------------------------------------------------------------------------------------")
                print(lens)
                f += 2; i = 0; j = 0; n = 2; colors = 0; string = str(""); string_2 = str(""); day = "None"; lens -= 1; day_week = 1;
            if lens == 0:
                f = 0;
                i = 0;
                j = 0;
                n = 0;
                colors = 0;
                break
    db.commit()
    db.close()
def db_f():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    cursor.execute('''SELECT * FROM timetable''')
    results = cursor.fetchall()
    for row in results:
        print(row)

def db_del():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    cursor.execute('''
            UPDATE timetable
            SET pair_number = ' ',
            teacher_name = ' ',
            classroom = ' '
            WHERE discipline IN ('ДЕНЬ', 'САМОСТОЯТЕЛЬНЫХ', 'ЗАНЯТИЙ');
    ''')
    db.commit()
    db.close()

def db_del_test():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    cursor.execute('''
            UPDATE timetable
            SET pair_number = ' ',
            teacher_name = ' ',
            classroom = ' '
            WHERE discipline IN ('ДЕНЬ', 'САМОСТОЯТЕЛЬНЫХ', 'ЗАНЯТИЙ');
    ''')
    db.commit()
    db.close()

def db_del_NIR():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    cursor.execute('''
            UPDATE timetable
            SET pair_number = ' ',
            teacher_name = ' ',
            classroom = ' '
            WHERE discipline IN ('НАУЧНО-', 'РАБОТА', 'ИССЛЕДОВАТЕЛЬСКАЯ');
    ''')
    db.commit()
    db.close()


def db_del_NIR_test():
    db = sqlite3.connect('sql/timetable_db.db')
    cursor = db.cursor()
    cursor.execute('''
            UPDATE timetable
            SET discipline = 'НАУЧНО-ИССЛЕДОВАТЕЛЬСКАЯ'
            WHERE discipline IN ('НАУЧНО-');
    ''')
    db.commit()
    db.close()


db_select()
db_start()
db_del_NIR()
db_del_NIR_test()
db_del_test()


# db_del()
db_f()
